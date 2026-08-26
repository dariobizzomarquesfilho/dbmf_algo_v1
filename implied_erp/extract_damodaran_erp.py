"""Extract country-level ERP data from Damodaran's ctryprem*.xls/.xlsx spreadsheet.

Layout-robust extractor: it discovers the worksheet, header row, and field→column
mapping dynamically (fuzzy token matching) instead of relying on fixed positions,
so it works across the modern .xlsx and legacy .xls archive files whose internal
layout differs (shifted header row, reordered/shuffled columns, renamed sheet).

Supports both modern .xlsx (openpyxl) and legacy .xls (xlrd) formats.

Usage:
    python extract_damodaran_erp.py --path "path/to/ctrypremJuly26.xlsx" --out "output.json"
    python extract_damodaran_erp.py --path "path/to/ctryprem00.xls" --out "output.json"
    python extract_damodaran_erp.py --path "path/to/ctrypremJuly26.xlsx"  # prints to stdout
    python extract_damodaran_erp.py --path "..." --report                # emit layout report to stderr
"""

from __future__ import annotations

import argparse
import difflib
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Iterator, Iterable

import openpyxl

# ── Constants ────────────────────────────────────────────────────────────────
# Default output path: <script_dir>/data/july26.json
DEFAULT_OUT = Path(__file__).resolve().parent / "data" / "july26.json"

# Fields we expect in the regular-country table, keyed by fuzzy label tokens.
# Expected label → (output field name, is_string)
REGULAR_FIELD_LABELS = [
    ("Region", "region", True),
    ("Total Equity Risk Premium", "total_equity_risk_premium", False),
    ("Country Risk Premium", "country_risk_premium", False),
    ("Moody's", "moody_rating", True),
    ("Rating", "moody_rating", True),
    ("Default Spread", "rating_default_spread", False),
    ("CDS", "sovereign_cds_net", False),
    ("Sovereign", "sovereign_cds_net", False),
    ("Total Equity Risk Premium 2", "total_equity_risk_premium2", False),
    ("Total Equity Risk Premium 3", "total_equity_risk_premium3", False),
    # Legacy (~2000-era) layouts: "Total Risk Premium" / "Adj. Default Spread"
    ("Total Risk Premium", "total_equity_risk_premium", False),
    ("Adj. Default Spread", "rating_default_spread", False),
    ("Long-Term Rating", "moody_rating", True),
]

# Frontier-market table expected labels.
FRONTIER_FIELD_LABELS = [
    ("PRS", "prs_score", False),
    ("ERP", "erp", False),
    ("Country Risk Premium", "crp", False),
    ("Default Spread", "default_spread", False),
]

# Metadata prompt phrases (substring, case-insensitive).
META_DATE_PROMPT = "date of update"
META_MATURE_PROMPT = "risk premium for a mature equity market"
META_US_PROMPT = "risk premium for the us"

# Country-name canonicalization (one stable key across all years).
COUNTRY_ALIASES = {
    "u.s.": "United States",
    "us": "United States",
    "usa": "United States",
    "united states of america": "United States",
    "russia": "Russian Federation",
    "korea": "South Korea",
    "uk": "United Kingdom",
    "britain": "United Kingdom",
    "germany": "Germany",
}

# How many top rows to scan for metadata prompts before the country table.
METADATA_MAX_ROW = 12


# ── Row-source abstraction ──────────────────────────────────────────────────

class _SheetRows:
    """Minimal row-source protocol for the extractor.

    Both backends expose .iter_rows(min_row, max_row, values_only=True)
    yielding tuples of cell values.  The abstraction hides engine-specific
    differences (0-based vs 1-based indexing, empty-cell representation).
    """

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = True,
    ) -> Iterator[tuple]:
        raise NotImplementedError

    @property
    def name(self) -> str:
        raise NotImplementedError

    @property
    def nrows(self) -> int:
        raise NotImplementedError

    @property
    def ncols(self) -> int:
        raise NotImplementedError


class _OpenpyxlSheet(_SheetRows):
    """Wraps an openpyxl worksheet."""

    def __init__(self, ws) -> None:
        self._ws = ws

    @property
    def name(self) -> str:
        return self._ws.title

    @property
    def nrows(self) -> int:
        return self._ws.max_row or 0

    @property
    def ncols(self) -> int:
        return self._ws.max_column or 0

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = True,
    ) -> Iterator[tuple]:
        return self._ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            values_only=values_only,
        )


class _XlrdSheet(_SheetRows):
    """Wraps an xlrd sheet.  Rows are 0-based; pads ragged rows to a uniform width."""

    def __init__(self, sheet, book) -> None:
        self._sheet = sheet
        self._book = book
        self._ncols = sheet.ncols

    @property
    def name(self) -> str:
        return self._sheet.name

    @property
    def nrows(self) -> int:
        return self._sheet.nrows

    @property
    def ncols(self) -> int:
        return self._sheet.ncols

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = True,
    ) -> Iterator[tuple]:
        # xlrd rows are 0-based; min_row/max_row are 1-based (Excel convention)
        lo = max(0, min_row - 1)
        hi = (
            self._sheet.nrows
            if max_row is None
            else min(max_row, self._sheet.nrows)
        )
        for r in range(lo, hi):
            vals = list(self._sheet.row_values(r))
            # Pad ragged rows so all tuples have the same length
            if len(vals) < self._ncols:
                vals += [None] * (self._ncols - len(vals))
            yield tuple(vals)


def _open_sheet(path: Path) -> tuple[_SheetRows, str, object]:
    """Dispatch by file extension; choose the worksheet by fuzzy title match.

    Falls back to the first sheet if no sheet name matches "erp" + "countr".
    Returns (sheet_rows, engine, book_or_None).
    """
    ext = path.suffix.lower()
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path))
        sheet = _pick_xlrd_sheet(book)
        return _XlrdSheet(sheet, book), "xlrd", book

    # Default: treat as .xlsx (openpyxl)
    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheet = _pick_openpyxl_sheet(wb)
    return _OpenpyxlSheet(sheet), "openpyxl", wb


def _pick_openpyxl_sheet(wb) -> object:
    names = list(wb.sheetnames)
    return wb[_pick_sheet_name(names)]


def _pick_xlrd_sheet(book) -> object:
    names = book.sheet_names()
    return book.sheet_by_name(_pick_sheet_name(names))


def _pick_sheet_name(names: list[str]) -> str:
    """Fuzzy-select a country-ERP sheet; fallback to the first sheet."""
    lowered = {n.lower(): n for n in names}
    for n_low, n in lowered.items():
        if "erp" in n_low and "countr" in n_low:
            return n
    # looser: any sheet mentioning 'country' or 'erp'
    for token in ("country", "erp"):
        for n_low, n in lowered.items():
            if token in n_low:
                return n
    return names[0]


# ── Date parsing ───────────────────────────────────────────────────────────

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_human_date(text: str) -> str | None:
    """Parse human-readable dates like 'July 9, 2026' or '9 July 2026'."""
    if not text:
        return None
    t = text.strip()
    # "Month DD, YYYY"
    m = re.search(r"([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(\d{4})", t)
    if m:
        mon = _MONTH_MAP.get(m.group(1).lower()[:3])
        if mon:
            try:
                return datetime(int(m.group(3)), mon, int(m.group(2))).strftime("%Y-%m-%d")
            except ValueError:
                return None
    # "DD Month YYYY"
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})", t)
    if m:
        mon = _MONTH_MAP.get(m.group(2).lower()[:3])
        if mon:
            try:
                return datetime(int(m.group(3)), mon, int(m.group(1))).strftime("%Y-%m-%d")
            except ValueError:
                return None
    return None


def _cell_to_date_str(value, book=None) -> str | None:
    """Convert a date cell value to 'YYYY-MM-DD'.

    Handles openpyxl datetime objects, xlrd float/int serial dates, and
    human-readable text.  Returns None if unparseable.
    """
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        # openpyxl sometimes yields "2026-07-09 00:00:00" or with microseconds
        s = value.strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}.*", s):
            s = s.split(" ")[0].split("T")[0]
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
                return s
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
            return s
        # xlrd sometimes stores the date serial as text, e.g. "44196.0" or "44196.00"
        # Strip trailing .0 variants robustly
        sv = s
        # Remove trailing .0+ (e.g. 44196.00 -> 44196)
        if "." in sv:
            try:
                fv = float(sv)
                if fv == int(fv):
                    sv = str(int(fv))
                else:
                    # Fractional serial (half-day) — truncate to int days
                    sv = str(int(fv))
            except ValueError:
                pass
        if sv.lstrip("-").isdigit() and 30000 < int(sv) < 60000:
            from datetime import timedelta as _td

            return (datetime(1899, 12, 30) + _td(days=int(sv))).strftime("%Y-%m-%d")
        return _parse_human_date(s)
    if isinstance(value, (int, float)) and book is not None:
        try:
            import xlrd as _xlrd

            # xlrd datemode is only on xlrd Book; openpyxl Workbook has no datemode
            datemode = getattr(book, "datemode", 0)
            y, mo, d, _, _, _ = _xlrd.xldate.xldate_as_tuple(float(value), datemode)
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except Exception:
            return None
    return None


# ── Helpers ─────────────────────────────────────────────────────────────────

def _to_float_or_none(value) -> float | None:
    """Convert value to float, returning None for NA/#N/A/empty."""
    if value is None:
        return None
    if isinstance(value, str) and value.strip().upper() in ("NA", "#N/A", "N/A", ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_token(label: str) -> str:
    """Lowercase + collapse whitespace for fuzzy matching."""
    return re.sub(r"\s+", " ", str(label).strip().lower()) if label is not None else ""


def _canonical_country(name: str) -> str:
    key = name.strip().lower()
    return COUNTRY_ALIASES.get(key, name.strip())


def _cell_text(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


# ── Layout detection ────────────────────────────────────────────────────────

def _detect_layout(rows: _SheetRows) -> dict:
    """Scan all rows to discover header row, field→column map, frontier section.

    Returns a dict with:
      sheet_name, header_row (1-based), field_col_map {field: col},
      frontier_header_row (1-based or None), frontier_col_map {field: col},
      metadata {prompt: (row, value_col)}, best_header_score.
    """
    ncols = max(rows.ncols, 1)
    # Build a normalized list of header candidate rows: rows whose first cell
    # contains 'country' (case-insensitive) OR a row containing many expected tokens.
    expected_tokens = [lab for lab, _, _ in REGULAR_FIELD_LABELS]
    candidate_header_row = None
    best_score = 0.0
    frontier_header_row = None
    frontier_best_score = 0.0

    all_rows = list(rows.iter_rows(min_row=1, values_only=True))

    for i, row in enumerate(all_rows, start=1):
        # Skip fully empty rows
        if not any(c is not None and str(c).strip() for c in row):
            continue
        # Stop scanning for headers past a reasonable depth.
        if i > METADATA_MAX_ROW + 60:
            break

        norm_cells = [_normalize_token(c) for c in row]
        # Heuristic: a header row mentions 'country' and several ERP-ish tokens.
        # Require the FIRST cell to be the 'Country' label so prose rows that
        # merely contain the word "country" (e.g. explanatory paragraphs) are
        # never mistaken for a table header.
        joined = " ".join(norm_cells)
        if norm_cells and norm_cells[0].startswith("country") and "country" in joined:
            score = _header_score(norm_cells, expected_tokens)
            if score > best_score:
                best_score = score
                candidate_header_row = i

        # Frontier header: mentions 'prs' and 'erp' / frontier section header.
        if "frontier" in joined and frontier_header_row is None:
            # The frontier *section* header row; real column header is usually the next row.
            # Look at the next non-empty row for the frontier column header.
            for j in range(i, min(i + 3, len(all_rows))):
                nxt = all_rows[j]
                if not any(c is not None and str(c).strip() for c in nxt):
                    continue
                nxt_norm = [_normalize_token(c) for c in nxt]
                fscore = _header_score(nxt_norm, [lab for lab, _, _ in FRONTIER_FIELD_LABELS])
                if fscore > frontier_best_score:
                    frontier_best_score = fscore
                    frontier_header_row = j + 1
                break

    field_col_map: dict[str, int] = {}
    frontier_col_map: dict[str, int] = {}

    if candidate_header_row is not None:
        header_cells = [_normalize_token(c) for c in all_rows[candidate_header_row - 1]]
        field_col_map = _match_fields(header_cells, REGULAR_FIELD_LABELS, ncols)

    if frontier_header_row is not None:
        fheader_cells = [_normalize_token(c) for c in all_rows[frontier_header_row - 1]]
        frontier_col_map = _match_fields(fheader_cells, FRONTIER_FIELD_LABELS, ncols)

    metadata = _detect_metadata(all_rows)

    return {
        "sheet_name": rows.name,
        "header_row": candidate_header_row,
        "field_col_map": field_col_map,
        "frontier_header_row": frontier_header_row,
        "frontier_col_map": frontier_col_map,
        "metadata": metadata,
        "best_header_score": round(best_score, 3),
    }


def _header_score(norm_cells: list[str], expected_tokens: list[str]) -> float:
    """Fraction of expected tokens that have a close match among cells."""
    if not norm_cells:
        return 0.0
    joined = " ".join(norm_cells)
    hits = 0
    for tok in expected_tokens:
        if tok.lower() in joined:
            hits += 1
    return hits / len(expected_tokens)


def _match_fields(
    header_cells: list[str], field_labels: list[tuple[str, str, bool]], ncols: int
) -> dict[str, tuple[int, bool]]:
    """Map output field names to (col_index, is_string) via fuzzy token matching.

    For each expected label, find the header cell whose normalized text is a
    close match (substring or difflib ratio).  Prefers an exact substring match.
    """
    col_map: dict[str, tuple[int, bool]] = {}
    used_cols: set[int] = set()
    # Collect candidate cells (non-empty, not already used).
    candidates = []
    for ci, cell in enumerate(header_cells):
        if ci >= ncols:
            break
        if not cell:
            continue
        candidates.append((ci, cell))

    for label, field, is_str in field_labels:
        lab_norm = _normalize_token(label)
        # 1) substring match (e.g. "Total Equity Risk Premium 2" in header)
        best_ci = None
        best_ratio = 0.0
        for ci, cell in candidates:
            if ci in used_cols:
                continue
            if lab_norm in cell or cell in lab_norm:
                # Prefer the closest-length match to avoid "ERP" matching "Country Risk Premium"
                ratio = difflib.SequenceMatcher(None, lab_norm, cell).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_ci = ci
        if best_ci is not None:
            col_map[field] = (best_ci, is_str)
            used_cols.add(best_ci)
            continue
        # 2) fuzzy match via difflib
        for ci, cell in candidates:
            if ci in used_cols:
                continue
            ratio = difflib.SequenceMatcher(None, lab_norm, cell).ratio()
            if ratio > best_ratio and ratio >= 0.6:
                best_ratio = ratio
                best_ci = ci
        if best_ci is not None:
            col_map[field] = (best_ci, is_str)
            used_cols.add(best_ci)
    return col_map


def _detect_metadata(all_rows: list[tuple]) -> dict[str, tuple[int, int]]:
    """Find metadata prompt rows in the leading rows.

    Returns {prompt_key: (row_1based, value_col_index)} where value_col_index is
    the first non-empty cell after col 0 in that row (or col 1).
    """
    meta: dict[str, tuple[int, int]] = {}
    for i, row in enumerate(all_rows, start=1):
        if i > METADATA_MAX_ROW:
            break
        first = _cell_text(row, 0)
        if not first:
            continue
        low = first.lower()
        for prompt, key in (
            (META_DATE_PROMPT, "date"),
            (META_MATURE_PROMPT, "mature"),
            (META_US_PROMPT, "us"),
        ):
            if prompt in low:
                # value column = first non-empty beyond col 0, else col 1
                val_col = 1
                for ci in range(1, len(row)):
                    if row[ci] is not None and str(row[ci]).strip():
                        val_col = ci
                        break
                meta[key] = (i, val_col)
                break
    return meta


# ── Extraction ──────────────────────────────────────────────────────────────

def _extract_regular_countries(rows: _SheetRows, layout: dict) -> tuple[dict, int]:
    """Extract regular (rated) countries using detected field_col_map."""
    countries: dict[str, dict] = {}
    header_row = layout["header_row"]
    col_map = layout["field_col_map"]
    if header_row is None:
        return countries, 0

    for i, row in enumerate(rows.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        first_col = row[0] if len(row) > 0 else None

        # Empty first-col is not necessarily footer — could be a blank
        # separator row. Skip it instead of aborting (footer is detected by
        # consecutive empties or reaching end, but a single blank should not
        # truncate the table).
        if first_col is None or (isinstance(first_col, str) and not first_col.strip()):
            continue

        # Stop when we hit the frontier markets section (case-insensitive)
        if isinstance(first_col, str) and "frontier markets" in first_col.lower():
            break

        # Skip non-country rows (e.g. the frontier header is separate; section header handled above)
        if not isinstance(first_col, str):
            continue

        country_name = _canonical_country(first_col)
        entry = {"is_frontier": False}
        for field, (col, is_str) in col_map.items():
            if col < len(row):
                raw = row[col]
                if is_str:
                    if raw is None:
                        entry[field] = None
                    elif isinstance(raw, str) and raw.strip().upper() in ("NA", "#N/A", "N/A", ""):
                        entry[field] = None
                    else:
                        entry[field] = str(raw).strip()
                else:
                    entry[field] = _to_float_or_none(raw)
            else:
                entry[field] = None
        # De-dupe: keep first occurrence if name repeats.
        if country_name not in countries:
            countries[country_name] = entry

    return countries, len(countries)


def _extract_frontier_countries(rows: _SheetRows, layout: dict) -> tuple[dict, int]:
    """Extract frontier markets (after the 'Frontier Markets' section header)."""
    countries: dict[str, dict] = {}
    fheader_row = layout.get("frontier_header_row")
    col_map = layout.get("frontier_col_map")
    if fheader_row is None or not col_map:
        return countries, 0

    started = False
    for i, row in enumerate(rows.iter_rows(min_row=1, values_only=True), start=1):
        first_col = row[0] if len(row) > 0 else None

        # Detect the frontier markets section header (case-insensitive)
        if isinstance(first_col, str) and "frontier markets" in first_col.lower():
            started = True
            continue
        if not started:
            continue

        # Skip until we pass the detected frontier header row
        if i < fheader_row:
            continue

        # Empty row in frontier section — skip, don't break (blank separator)
        if first_col is None or (isinstance(first_col, str) and not first_col.strip()):
            continue

        if not isinstance(first_col, str):
            continue

        country_name = _canonical_country(first_col)
        entry = {"is_frontier": True}
        for field, (col, _is_str) in col_map.items():
            if col < len(row):
                entry[field] = _to_float_or_none(row[col])
            else:
                entry[field] = None
        if country_name not in countries:
            countries[country_name] = entry

    return countries, len(countries)


def _extract_metadata_values(rows: _SheetRows, layout: dict, book=None) -> tuple[str | None, float | None, float | None, str]:
    """Resolve metadata (updated date, mature_market_erp, us_erp) + its source tag."""
    meta = layout["metadata"]
    mature_market_erp = None
    us_erp = None
    updated = None
    date_source = "none"

    all_rows = list(rows.iter_rows(min_row=1, max_row=METADATA_MAX_ROW, values_only=True))

    if "date" in meta:
        r, c = meta["date"]
        val = all_rows[r - 1][c] if r - 1 < len(all_rows) and c < len(all_rows[r - 1]) else None
        if val is not None:
            ds = _cell_to_date_str(val, book)
            if ds:
                updated = ds
                date_source = "cell"
            else:
                updated = str(val).strip()
                date_source = "cell-text"

    if "mature" in meta:
        r, c = meta["mature"]
        if r - 1 < len(all_rows) and c < len(all_rows[r - 1]):
            mature_market_erp = _to_float_or_none(all_rows[r - 1][c])

    if "us" in meta:
        r, c = meta["us"]
        if r - 1 < len(all_rows) and c < len(all_rows[r - 1]):
            us_erp = _to_float_or_none(all_rows[r - 1][c])

    return updated, mature_market_erp, us_erp, date_source


# ── Public API ──────────────────────────────────────────────────────────────

def extract(path: str) -> dict:
    """Read the spreadsheet and return structured JSON-compatible dictionary.

    Works with both .xlsx (openpyxl) and .xls (xlrd) formats.  Layout is
    discovered dynamically; output schema is identical to the legacy extractor.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    rows, engine, book = _open_sheet(p)

    layout = _detect_layout(rows)

    updated, mature_market_erp, us_erp, date_source = _extract_metadata_values(rows, layout, book=book)

    # Extract both sections
    countries, n_regular = _extract_regular_countries(rows, layout)
    frontier, n_frontier = _extract_frontier_countries(rows, layout)
    countries.update(frontier)

    if not countries:
        raise ValueError(f"Nenhum país encontrado em {path} (layout: {layout.get('sheet_name')})")

    data = {
        "source": p.name,
        "updated": updated,
        "mature_market_erp": mature_market_erp,
        "us_erp": us_erp,
        "countries": countries,
    }
    # Stash layout diagnostics (not part of the canonical schema; consumed by --report).
    data["_layout"] = {
        "sheet": layout["sheet_name"],
        "engine": engine,
        "header_row": layout["header_row"],
        "header_score": layout["best_header_score"],
        "field_col_map": layout["field_col_map"],
        "frontier_header_row": layout["frontier_header_row"],
        "frontier_col_map": layout["frontier_col_map"],
        "date_source": date_source,
        "n_regular": n_regular,
        "n_frontier": n_frontier,
    }
    return data


def build_layout_report(data: dict) -> str:
    """Render a human-readable layout report (for --report / stderr)."""
    lay = data.get("_layout", {})
    lines = [
        f"[{data.get('source')}] engine={lay.get('engine')} sheet='{lay.get('sheet')}'",
        f"  header_row={lay.get('header_row')} (match_score={lay.get('header_score')})",
        f"  field_col_map={lay.get('field_col_map')}",
        f"  frontier_header_row={lay.get('frontier_header_row')} frontier_col_map={lay.get('frontier_col_map')}",
        f"  countries: regular={lay.get('n_regular')} frontier={lay.get('n_frontier')} total={len(data.get('countries', {}))}",
        f"  updated={data.get('updated')!r} (source={lay.get('date_source')}) "
        f"us_erp={data.get('us_erp')} mature_market_erp={data.get('mature_market_erp')}",
    ]
    score = lay.get("header_score", 1.0)
    total = len(data.get("countries", {}))
    if score < 0.5 or total == 0:
        lines.append("  !! WARNING: low header match score or 0 countries — review extraction!")
    return "\n".join(lines)


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Extrai dados de ERP por país da planilha Damodaran ctryprem*.xls/.xlsx (layout-robust)"
    )
    ap.add_argument(
        "--path",
        "--xlsx",
        dest="path",
        required=True,
        help="Caminho para o arquivo ctryprem*.xls ou ctryprem*.xlsx",
    )
    ap.add_argument(
        "--out",
        default=str(DEFAULT_OUT),
        help="Caminho do arquivo JSON de saída (padrão: data/july26.json)",
    )
    ap.add_argument(
        "--report",
        action="store_true",
        help="Imprimir relatório de layout para stderr",
    )
    args = ap.parse_args()

    data = extract(args.path)

    if args.report:
        print(build_layout_report(data), file=__import__("sys").stderr)

    json_str = json.dumps(data, indent=2, ensure_ascii=False)

    if args.out:
        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json_str, encoding="utf-8")
        print(f"[OK] Dados salvos em: {out_path}")
        print(f"     Países: {len(data['countries'])}")
        print(f"     Fonte:  {data['source']}")
    else:
        print(json_str)


if __name__ == "__main__":
    main()
