"""Tests for the layout-robust Damodaran ERP extractor.

Builds synthetic in-memory openpyxl sheets whose header row / column order /
sheet name differ from the current file, then asserts extract() still recovers
countries and metadata.  No Lean imports, no network.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

# Make implied_erp importable when run from this dir.
REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from implied_erp.extract_damodaran_erp import extract  # noqa: E402


def _write_sheet(tmp_path: Path, build_fn) -> str:
    wb = openpyxl.Workbook()
    build_fn(wb)
    out = tmp_path / "ctryprem_test.xlsx"
    wb.save(out)
    return str(out)


def test_shifted_header_and_shuffled_columns(tmp_path):
    """Header row pushed down; columns reordered; sheet renamed."""

    def build(wb):
        ws = wb.active
        ws.title = "Country ERPs (renamed)"
        # Metadata in rows 1-3, header at row 5, data from row 6.
        ws["A1"] = "Date of update:"
        ws["B1"] = "2024-07-09"
        ws["A2"] = "Enter the current risk premium for a mature equity market"
        ws["E2"] = 0.0423
        ws["A3"] = "Enter the current risk premium for the US"
        ws["E3"] = 0.0440
        # Shuffled header order: ERP, Country, CRP, Rating
        ws["A5"] = "Country"
        ws["B5"] = "Total Equity Risk Premium"
        ws["C5"] = "Country Risk Premium"
        ws["D5"] = "Moody's Rating"
        ws["A6"] = "United States"
        ws["B6"] = 0.0440
        ws["C6"] = 0.0
        ws["D6"] = "Aaa"
        ws["A7"] = "Brazil"
        ws["B7"] = 0.0725
        ws["C7"] = 0.0350
        ws["D7"] = "Ba2"

    data = extract(_write_sheet(tmp_path, build))
    assert data["updated"] == "2024-07-09"
    assert data["mature_market_erp"] == 0.0423
    assert data["us_erp"] == 0.0440
    countries = data["countries"]
    assert "United States" in countries and "Brazil" in countries
    assert countries["United States"]["total_equity_risk_premium"] == 0.0440
    assert countries["United States"]["moody_rating"] == "Aaa"
    assert countries["Brazil"]["country_risk_premium"] == 0.0350


def test_first_sheet_fallback_when_name_drifts(tmp_path):
    """Sheet name has no 'country'/'erp' token — must fall back to first sheet."""

    def build(wb):
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "Date of update:"
        ws["B1"] = "July 9, 2024"  # human-readable date
        ws["A3"] = "Country"
        ws["B3"] = "Total Equity Risk Premium"
        ws["A4"] = "Germany"
        ws["B4"] = 0.0550
        ws["A5"] = "Japan"
        ws["B5"] = 0.0610

    data = extract(_write_sheet(tmp_path, build))
    # Human-readable date parsed via month map.
    assert data["updated"] == "2024-07-09"
    assert "Germany" in data["countries"]
    assert data["countries"]["Germany"]["total_equity_risk_premium"] == 0.0550


def test_no_countries_raises(tmp_path):
    def build(wb):
        ws = wb.active
        ws.title = "ERPs by country"
        ws["A1"] = "Date of update:"
        ws["B1"] = "2024-07-09"

    import pytest

    with pytest.raises(ValueError):
        extract(_write_sheet(tmp_path, build))


def test_country_alias_canonicalization(tmp_path):
    """'U.S.' should be canonicalized to 'United States'."""

    def build(wb):
        ws = wb.active
        ws.title = "ERPs by country"
        ws["A1"] = "Date of update:"
        ws["B1"] = "2024-07-09"
        ws["A3"] = "Country"
        ws["B3"] = "Total Equity Risk Premium"
        ws["A4"] = "U.S."
        ws["B4"] = 0.0440

    data = extract(_write_sheet(tmp_path, build))
    assert "United States" in data["countries"]
    assert "U.S." not in data["countries"]
